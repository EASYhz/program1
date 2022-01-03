#BTC 마켓은 나중에 업뎃

from upbit.client import Upbit
from openpyxl import Workbook
from datetime import datetime

now = datetime.now()
current_day = str(now.year) + '.' + str(now.month) + '.' + str(now.day)
current_time = now.strftime("%H:%M:%S")



access_key = "OHIHQamGBbrAkb4WS4uQASRKcB8dhIILup38U3L1"
secret_key = "4equVuojIxcV0QxxnJkoOFnFBKJaWZxQFCCKfbay"

client = Upbit(access_key, secret_key)

#resp = client.Market.Market_info_all()
marketList = client.Market.Market_info_all()

ticker = []
for i in range(0, len(marketList['result'])):
    ticker.append(marketList['result'][i]['market'])        #이름 나열
    
str_match = [s for s in ticker if "KRW" in s]
#print(str_match)

tickerListAdd = ", ".join(str_match)
'''
for i in range(0, int(len(str_match)/2)):
    price = client.Candle.Candle_minutes(
        unit=10,
        market=str_match[i]
    )
    print(price['result'])
for i in range(int(len(str_match)/2), len(str_match)):
    price = client.Candle.Candle_minutes(
        unit=10,
        market=str_match[i]
    )
    print(price['result'])
'''
resp = client.Trade.Trade_ticker(
    markets= tickerListAdd
)
name =[]
tradePrice24 =[]
for i in range(0, len(str_match)):
    name.append(resp['result'][i]['market'])
    tradePrice24.append(int(resp['result'][i]['acc_trade_price_24h']/1000000))
    


write_wb = Workbook()
write_ws = write_wb.active
write_ws['A1'] = '날짜'
write_ws['B1'] = current_day
write_ws['A2'] = '시간'
write_ws['B2'] = current_time
write_ws['A3'] = '종목'
write_ws['B3'] = '거래대금'

for i in range(0, len(str_match)):
    write_ws.cell(4+i, 1).value = name[i]
    write_ws.cell(4+i, 2).value = tradePrice24[i]
    
write_wb.save('C:\\Users\\LG\\Desktop\\업비트 거래대금 크롤링\\update111111.xlsx')